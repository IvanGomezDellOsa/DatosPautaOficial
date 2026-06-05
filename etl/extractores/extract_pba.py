# -*- coding: utf-8 -*-
"""
extract_pba.py  —  Extrae y normaliza Provincia de Buenos Aires (PBA).

Fuentes:
  A) Tabula CSV 2020-2024  (PBA Nuevos datos sin curar/YYYY/YYYY CSV/tabula-*.csv)
     Cada archivo = uno o varios expedientes mensuales:
        Expediente N° 2140-3166/2020,,,
        Orden de Publicidad,Proveedor,Medio,Monto
        832,TV Y FM LAS BRISAS S.R.L.,FM 98.5 ...,"$ 9.457,00"
     - fila Expediente -> resolucion (se arrastra para las filas siguientes)
     - se descartan filas donde "Orden" no es numerico (subtotales / headers repetidos)
     - fecha: del mes del NOMBRE del archivo (tabula-Abril 2020 -> 2020-04-01)
     - NO se usa PBA_*_sin_curar_unificado.csv (duplicaria los tabula-*)

  B) Excel 2025  (Datos crudos/PBA/Publicidad Oficial de PBA - 2025.xlsx)
     Columnas: Operación, Proveedor, Medio, Grupo, Importe, Fecha documento GEDO,
               Resolución, Resumen, Link, Mes asignado en resumen, Monto deflactado
     - monto = Importe ; fecha = Mes asignado ; resolucion = Resolución
"""

import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import comun  # noqa: E402
from comun import (  # noqa: E402
    DIR_PBA, DIR_PBA_NUEVOS, registro, normalizar_monto, normalizar_fecha,
    limpiar_str, leer_filas, detectar_sep, MESES, Contador, log,
)

JURIS = "PBA"


def _fecha_de_nombre(nombre):
    """Devuelve (anio, fecha_iso|None) a partir del nombre del archivo tabula."""
    n = nombre.lower()
    anio = comun.anio_de_texto(n)
    mes = None
    for nombre_mes, num in MESES.items():
        if len(nombre_mes) > 3 and re.search(r"\b" + nombre_mes + r"\b", n):
            mes = num
            break
    if anio and mes:
        return anio, f"{anio:04d}-{mes:02d}-01"
    return anio, None


def _parse_tabula(path, cont):
    out = []
    anio, fecha = _fecha_de_nombre(path.name)
    if anio is None:
        # fallback: año de la carpeta padre
        anio = comun.anio_de_texto(str(path.parent))
    sep = detectar_sep(path)
    filas = leer_filas(path, sep)

    expediente = None
    idx = None  # indices de columnas una vez detectado el header
    for row in filas:
        celdas = [(c or "").strip() for c in row]
        linea = " ".join(celdas)
        # fila de expediente
        m = re.search(r"[Ee]xpediente\s*N[°ºo:]*\s*([\w\-/.]+)", linea)
        if m and "proveedor" not in linea.lower():
            expediente = m.group(1)
            idx = None
            continue
        # fila de header
        low = [c.lower() for c in celdas]
        if any("proveedor" in c for c in low) and any("medio" in c for c in low):
            idx = {"orden": None, "prov": None, "medio": None, "monto": None}
            for i, c in enumerate(low):
                if idx["orden"] is None and ("orden" in c or c == "op" or c.startswith("op ") or c == "op."):
                    idx["orden"] = i
                if idx["prov"] is None and "proveedor" in c:
                    idx["prov"] = i
                if idx["medio"] is None and "medio" in c:
                    idx["medio"] = i
                if idx["monto"] is None and ("monto" in c or "importe" in c):
                    idx["monto"] = i
            if idx["orden"] is None:
                idx["orden"] = 0   # por defecto, primera columna
            continue
        if idx is None:
            continue
        # fila de datos: "Orden" debe ser numerico
        orden_val = celdas[idx["orden"]] if idx["orden"] is not None and idx["orden"] < len(celdas) else ""
        if not re.fullmatch(r"\d+", orden_val.strip()):
            cont.descartar("orden_no_numerico")
            continue
        monto = normalizar_monto(celdas[idx["monto"]]) if idx["monto"] is not None and idx["monto"] < len(celdas) else None
        if monto is None:
            cont.descartar("monto_invalido")
            continue
        prov = limpiar_str(celdas[idx["prov"]]) if idx["prov"] is not None and idx["prov"] < len(celdas) else None
        medio = limpiar_str(celdas[idx["medio"]]) if idx["medio"] is not None and idx["medio"] < len(celdas) else None
        if prov is None and medio is None:
            cont.descartar("sin_prov_ni_medio")
            continue
        out.append(registro(
            JURIS, anio, monto, path.name,
            fecha=fecha, proveedor=prov, medio=medio, resolucion=expediente))
        cont.ok()
    return out


def _parse_excel_2025(path, cont):
    out = []
    try:
        import openpyxl
    except ImportError:
        log("[PBA] ADVERTENCIA: openpyxl no disponible; se omite el Excel 2025")
        cont.descartar("sin_openpyxl")
        return out
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = ws.iter_rows(values_only=True)
    header = next(filas, None)
    if not header:
        return out
    H = [(str(h).lower().strip() if h is not None else "") for h in header]

    def col(*claves):
        for k in claves:
            for i, h in enumerate(H):
                if k == h:
                    return i
        for k in claves:
            for i, h in enumerate(H):
                if k in h:
                    return i
        return None

    i_prov = col("proveedor")
    i_medio = col("medio")
    i_monto = col("importe", "monto")
    i_res = col("resolución", "resolucion")
    i_fecha = col("mes asignado", "fecha documento gedo", "fecha")
    for row in filas:
        if row is None:
            continue
        val = row[i_monto] if i_monto is not None and i_monto < len(row) else None
        monto = normalizar_monto(val) if not isinstance(val, (int, float)) else (abs(float(val)) if val else None)
        if monto is None or monto == 0:
            cont.descartar("monto_invalido")
            continue
        fval = row[i_fecha] if i_fecha is not None and i_fecha < len(row) else None
        fecha = None
        anio = 2025
        if hasattr(fval, "year"):
            fecha = f"{fval.year:04d}-{fval.month:02d}-{fval.day:02d}"
            anio = fval.year
        elif fval:
            fecha = normalizar_fecha(str(fval), 2025)
            anio = comun.anio_de_texto(str(fval)) or 2025
        prov = limpiar_str(str(row[i_prov])) if i_prov is not None and row[i_prov] is not None else None
        medio = limpiar_str(str(row[i_medio])) if i_medio is not None and row[i_medio] is not None else None
        if prov is None and medio is None:
            cont.descartar("sin_prov_ni_medio")
            continue
        res = row[i_res] if i_res is not None and i_res < len(row) else None
        out.append(registro(
            JURIS, anio, monto, path.name,
            fecha=fecha, proveedor=prov, medio=medio,
            resolucion=limpiar_str(str(res)) if res is not None else None))
        cont.ok()
    return out


def extraer():
    log("[PBA] iniciando extraccion")
    registros = []
    cont = Contador("[PBA]")

    # A) tabula 2020-2024
    if DIR_PBA_NUEVOS.exists():
        tabulas = sorted(DIR_PBA_NUEVOS.glob("**/tabula-*.csv"))
        for f in tabulas:
            registros += _parse_tabula(f, cont)
    else:
        log(f"[PBA] ADVERTENCIA: no existe {DIR_PBA_NUEVOS}")

    # B) excel 2025
    if DIR_PBA.exists():
        for f in sorted(DIR_PBA.glob("*.xlsx")):
            registros += _parse_excel_2025(f, cont)

    cont.resumen()
    log(f"[PBA] total: {len(registros)} registros")
    return registros


if __name__ == "__main__":
    regs = extraer()
    print(f"\nPBA -> {len(regs)} registros")
    for r in regs[:3]:
        print(r)
    from collections import Counter
    print("por anio:", dict(sorted(Counter(r["anio"] for r in regs).items())))
    print("con fecha:", sum(1 for r in regs if r["fecha"]))
    print("con resolucion:", sum(1 for r in regs if r["resolucion"]))
